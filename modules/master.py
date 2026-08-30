import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components
import os

FILE_MASTER_COA = "master_coa_bss.xlsx"
FILE_MASTER_BU = "master_bu_bss.xlsx"
FILE_MASTER_PELANGGAN = "master_pelanggan_bss.xlsx"
FILE_MASTER_PEMASOK = "master_pemasok_bss.xlsx"
FILE_MASTER_GUDANG = "master_gudang_bss.xlsx"
FILE_MASTER_ALAT = "master_alat_bss.xlsx"

def simpan_excel_cantik(df, filepath, sheet_name='Sheet1'):
    """Fungsi otomatis menyimpan dataframe ke Excel dengan format rapi dan bergaris"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="000000")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 25
    
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="left" if col > 1 else "center")
            
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    wb.save(filepath)

def generate_kategori_otomatis(kode_str):
    kode_bersih = str(kode_str).strip()
    kategori_1, kategori_2, sub_kategori = "", "", ""
    if not kode_bersih or kode_bersih == "nan":
        return "", "", ""

    p1 = kode_bersih[0] if len(kode_bersih) > 0 else ""
    if p1 == '1': kategori_1 = "1 - Aktiva"
    elif p1 == '2': kategori_1 = "2 - Hutang"
    elif p1 == '3': kategori_1 = "3 - Ekuitas"
    elif p1 == '4': kategori_1 = "4 - Pendapatan"
    elif p1 == '5': kategori_1 = "5 - Harga Pokok Penjualan"
    elif p1 == '6': kategori_1 = "6 - Beban Usaha"
    elif p1 == '7': kategori_1 = "7 - Pendapatan / Beban Lain-Lain"
    elif p1 == '8': kategori_1 = "8 - Ekuitas / Deviden"
    elif p1 == '9': kategori_1 = "9 - Beban / Gain-Loss"
    
    p2 = kode_bersih[:2] if len(kode_bersih) >= 2 else ""
    if p2 == '11': kategori_2 = "11 - Aktiva Lancar"
    elif p2 == '12': kategori_2 = "12 - Aktiva Tetap / Lainnya"
    elif p2 == '13': kategori_2 = "13 - Aktiva Tetap / Akumulasi"
    elif p2 == '21': kategori_2 = "21 - Hutang Lancar"
    elif p2 == '22': kategori_2 = "22 - Hutang Jangka Panjang"
    elif p2 == '31': kategori_2 = "31 - Modal"
    elif p2 == '32': kategori_2 = "32 - Laba Ditahan"
    elif p2 == '41': kategori_2 = "41 - Pendapatan Proyek GS"
    elif p2 == '42': kategori_2 = "42 - Pendapatan Proyek CR"
    elif p2 == '51': kategori_2 = "51 - Harga Pokok Proyek GS"
    elif p2 == '52': kategori_2 = "52 - Harga Pokok Proyek CR"
    elif p2 == '53': kategori_2 = "53 - Harga Pokok Proyek Bengkel"
    elif p2 == '62': kategori_2 = "62 - Beban Administrasi"
    elif p2 == '71': kategori_2 = "71 - Pendapatan Lain-lain"
    elif p2 == '72': kategori_2 = "72 - Biaya Lain-lain"
    elif p2 == '75': kategori_2 = "75 - Pajak Penghasilan"
    elif p2 == '88': kategori_2 = "88 - Deviden"
    elif p2 == '91': kategori_2 = "91 - Gain or Loss IDR"

    p3 = kode_bersih[:3] if len(kode_bersih) >= 3 else ""
    if p3 == '111': sub_kategori = "111 - Kas"
    elif p3 == '112': sub_kategori = "112 - Bank"
    elif p3 == '113': sub_kategori = "113 - Investasi Jk Pendek"
    elif p3 == '114': sub_kategori = "114 - Piutang Usaha"
    elif p3 == '115': sub_kategori = "115 - Uang Muka Pembelian"
    elif p3 == '116': sub_kategori = "116 - Piutang Lain"
    elif p3 == '117': sub_kategori = "117 - Persediaan Barang"
    elif p3 == '118': sub_kategori = "118 - Uang Muka Biaya / Pajak"
    elif p3 == '121': sub_kategori = "121 - Investasi Jangka Panjang"
    elif p3 == '131': sub_kategori = "131 - Harga Peroleh Aktiva Tetap"
    elif p3 == '132': sub_kategori = "132 - Akumulasi Penyusutan"
    elif p3 == '211': sub_kategori = "211 - Hutang Bank"
    elif p3 == '212': sub_kategori = "212 - Hutang Usaha"
    elif p3 == '213': sub_kategori = "213 - Pendapatan Diterima Dimuka"
    elif p3 == '214': sub_kategori = "214 - Hutang Pajak"
    elif p3 == '215': sub_kategori = "215 - Hutang Biaya"
    elif p3 == '216': sub_kategori = "216 - Hutang Lain-Lain"
    elif p3 == '221': sub_kategori = "221 - Kewajiban Jangka Panjang"
    elif p3 == '311': sub_kategori = "311 - Modal"
    elif p3 == '321': sub_kategori = "321 - Laba Ditahan"
    elif p3 == '411': sub_kategori = "411 - Penjualan Barang Dagangan"
    elif p3 == '412': sub_kategori = "412 - Pendapatan Jasa Tronton"
    elif p3 == '413': sub_kategori = "413 - Pendapatan Proyek Jasa Umum"
    elif p3 == '421': sub_kategori = "421 - Pendapatan Proyek Konstruksi"
    elif p3 == '511': sub_kategori = "511 - Harga Pokok Pengadaan Barang"
    elif p3 == '512': sub_kategori = "512 - Harga Pokok Operasional Tronton"
    elif p3 == '513': sub_kategori = "513 - Harga Pokok Proyek Jasa Umum"
    elif p3 == '514': sub_kategori = "514 - Biaya Lain-Lain Proyek"
    elif p3 == '621': sub_kategori = "621 - Beban Administrasi Karyawan"
    elif p3 == '622': sub_kategori = "622 - Suplies Kantor"
    elif p3 == '623': sub_kategori = "623 - Jasa Pihak Ketiga"
    elif p3 == '624': sub_kategori = "624 - Biaya Transportasi"
    elif p3 == '625': sub_kategori = "625 - Biaya Umum Kantor"
    elif p3 == '626': sub_kategori = "626 - Pemeliharaan Aktiva"
    elif p3 == '627': sub_kategori = "627 - Iuran, Pungutan & Sumbangan"
    elif p3 == '628': sub_kategori = "628 - Penyusutan Aktiva"
    elif p3 == '629': sub_kategori = "629 - Beban Adm Lain-Lain"
    elif p3 == '710': sub_kategori = "710 - Pendapatan Lain-lain"
    elif p3 == '720': sub_kategori = "720 - Biaya Lain-lain"
    elif p3 == '750': sub_kategori = "750 - Pajak Penghasilan"
    elif p3 == '880': sub_kategori = "880 - Deviden"
    elif p3 == '910': sub_kategori = "910 - Realize / Unrealize Gain or Loss IDR"

    return sub_kategori, kategori_2, kategori_1

def muat_atau_buat_file(filepath, kolom_default, sheet_name):
    """Fungsi memastikan file fisik Excel selalu ada dengan kolom standar yang valid."""
    if os.path.exists(filepath):
        try:
            df_raw = pd.read_excel(filepath, header=0)
            if not df_raw.empty and all(col in df_raw.columns for col in kolom_default):
                df_clean = pd.DataFrame()
                for col in kolom_default:
                    df_clean[col] = df_raw[col].fillna("").astype(str).str.strip()
                df_clean = df_clean[df_clean.iloc[:, 0] != ""]
                return df_clean.reset_index(drop=True)
        except Exception:
            pass
            
    df_kosong = pd.DataFrame(columns=kolom_default)
    simpan_excel_cantik(df_kosong, filepath, sheet_name)
    return df_kosong

def muat_data_master_bersih():
    if os.path.exists(FILE_MASTER_COA):
        try:
            df_raw = pd.read_excel(FILE_MASTER_COA, header=0)
            kolom_coa = ["Kode Akun", "Nama Akun", "Sub Account", "Sub Kategori", "Kategori"]
            if not df_raw.empty and all(col in df_raw.columns for col in kolom_coa):
                df_clean = pd.DataFrame()
                df_clean['Kode Akun'] = df_raw['Kode Akun'].fillna("").astype(str).str.strip()
                df_clean['Nama Akun'] = df_raw['Nama Akun'].fillna("Tanpa Nama").astype(str)
                
                sub_accs, sub_kats, kats = [], [], []
                for kode in df_clean['Kode Akun']:
                    s_acc, s_kat, kat = generate_kategori_otomatis(kode)
                    sub_accs.append(s_acc)
                    sub_kats.append(s_kat)
                    kats.append(kat)
                    
                df_clean['Sub Account'] = sub_accs
                df_clean['Sub Kategori'] = sub_kats
                df_clean['Kategori'] = kats
                
                df_clean = df_clean[df_clean['Kode Akun'] != ""]
                return df_clean.reset_index(drop=True)
        except Exception:
            pass
            
    df_kosong = pd.DataFrame(columns=["Kode Akun", "Nama Akun", "Sub Account", "Sub Kategori", "Kategori"])
    simpan_excel_cantik(df_kosong, FILE_MASTER_COA, 'Master_COA')
    return df_kosong

def render_preview_dan_cetak(df, judul_laporan, file_path, key_prefix):
    """Fungsi helper seragam untuk tombol Download, Preview Resmi, dan Cetak"""
    st.markdown("---")
    st.markdown(f"#### 📥 Ekspor & Cetak Data {judul_laporan}")
    col_ex1, col_ex2, col_ex3 = st.columns(3)
    
    with col_ex1:
        if os.path.exists(file_path):
            with open(file_path, "rb") as f:
                excel_bytes = f.read()
            st.download_button(
                label="📥 Download File Excel",
                data=excel_bytes,
                file_name=file_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_{key_prefix}"
            )
    with col_ex2:
        state_key = f"show_preview_{key_prefix}"
        if state_key not in st.session_state: 
            st.session_state[state_key] = False
        if st.button("👁️ Preview Dokumen Resmi", use_container_width=True, key=f"btn_prev_{key_prefix}"):
            st.session_state[state_key] = not st.session_state[state_key]
            st.rerun()
    with col_ex3:
        components.html("""
        <button onclick="parent.window.print();" style="
            width: 100%; background-color: #ffffff; color: #31333F; 
            padding: 0.5rem 0.75rem; border: 1px solid #d6d6d8; 
            border-radius: 0.3rem; cursor: pointer; font-weight: 500;
            font-family: Source Sans Pro, sans-serif; box-shadow: 0 1px 2px rgba(0,0,0,0.05);
        ">🖨️ Cetak / Print Dokumen</button>
        """, height=45)

    if st.session_state.get(f"show_preview_{key_prefix}", False):
        st.markdown("---")
        report_html = f"""
        <div id="printable-report" style="background-color: #ffffff; padding: 30px; border-radius: 8px; border: 1px solid #cccccc; color: #000000; font-family: Arial, sans-serif;">
            <h2 style="text-align: center; margin-bottom: 5px; color: #1F4E78;">PT BANGGAI SENTRAL SULAWESI</h2>
            <h4 style="text-align: center; margin-top: 0px; margin-bottom: 25px; color: #444444;">LAPORAN {judul_laporan.upper()}</h4>
            <hr style="border: 1px solid #1F4E78; margin-bottom: 20px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
                <thead>
                    <tr style="background-color: #1F4E78; color: white;">
        """
        for col_name in df.columns:
            report_html += f'<th style="border: 1px solid #b0b0b0; padding: 8px; text-align: center;">{col_name}</th>'
        report_html += "</tr></thead><tbody>"
        
        for _, row in df.iterrows():
            report_html += "<tr>"
            for col_name in df.columns:
                report_html += f'<td style="border: 1px solid #d0d0d0; padding: 6px; text-align: left;">{row[col_name]}</td>'
            report_html += "</tr>"
        report_html += "</tbody></table></div>"
        
        components.html(report_html, height=600, scrolling=True)
        if st.button("❌ Tutup Preview Dokumen", key=f"close_prev_{key_prefix}"):
            st.session_state[state_key] = False
            st.rerun()

def render_modul_0():
    # CSS untuk Sticky Header / Freeze Navigasi Master & Judul Modul di Atas Layar
    st.markdown("""
        <style>
            /* Freeze Header Utama & Navigasi Tab */
            .stTabs {
                position: sticky;
                top: 0px;
                z-index: 99999;
                background-color: #ffffff;
                padding-top: 10px;
                padding-bottom: 5px;
            }
            div[data-testid="stVerticalBlock"] > div:has(> div.stSubheader) {
                position: sticky;
                top: 0px;
                z-index: 99998;
                background-color: #ffffff;
            }
            @media print {
                body * { visibility: hidden; }
                #printable-report, #printable-report * { visibility: visible; }
                #printable-report { position: absolute; left: 0; top: 0; width: 100%; background: white !important; }
                [data-testid="stSidebar"], header, footer, .stButton { display: none !important; }
            }
        </style>
    """, unsafe_allow_html=True)

    # Inisialisasi & Validasi Ketat State Sesi & File Fisik
    if 'master_coa' not in st.session_state or st.session_state.master_coa.empty or 'Kode Akun' not in st.session_state.master_coa.columns:
        st.session_state.master_coa = muat_data_master_bersih()
    if 'master_bu' not in st.session_state or st.session_state.master_bu.empty or 'Kode BU' not in st.session_state.master_bu.columns:
        st.session_state.master_bu = muat_atau_buat_file(FILE_MASTER_BU, ["Kode BU", "Nama Business Unit", "Keterangan"], 'Master_BU')
    if 'master_pelanggan' not in st.session_state or st.session_state.master_pelanggan.empty or 'Kode Pelanggan' not in st.session_state.master_pelanggan.columns:
        st.session_state.master_pelanggan = muat_atau_buat_file(FILE_MASTER_PELANGGAN, ["Kode Pelanggan", "Nama Pelanggan", "Alamat", "Kontak"], 'Master_Pelanggan')
    if 'master_pemasok' not in st.session_state or st.session_state.master_pemasok.empty or 'Kode Pemasok' not in st.session_state.master_pemasok.columns:
        st.session_state.master_pemasok = muat_atau_buat_file(FILE_MASTER_PEMASOK, ["Kode Pemasok", "Nama Pemasok", "Alamat", "Kontak"], 'Master_Pemasok')
    if 'master_gudang' not in st.session_state or st.session_state.master_gudang.empty or 'Kode Gudang' not in st.session_state.master_gudang.columns:
        st.session_state.master_gudang = muat_atau_buat_file(FILE_MASTER_GUDANG, ["Kode Gudang", "Nama Gudang", "Lokasi", "Keterangan"], 'Master_Gudang')
    if 'master_alat' not in st.session_state or st.session_state.master_alat.empty or 'Kode Alat' not in st.session_state.master_alat.columns:
        st.session_state.master_alat = muat_atau_buat_file(FILE_MASTER_ALAT, ["Kode Alat", "Nama Alat / Unit", "Jenis / Kategori", "No. Polisi / Serial", "Keterangan"], 'Master_Alat')

    st.subheader("Modul 0: Pengaturan Master Data (COA, BU, Pelanggan, Pemasok, Gudang, & Alokasi Alat)")
    
    # Navigasi 6 Tab Utama
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📁 Master Akun (COA)", 
        "🏢 Business Unit (BU)", 
        "👥 Nama Pelanggan", 
        "🚚 Nama Pemasok", 
        "🏢 Gudang",
        "🚜 Alokasi Alat / Unit"
    ])
    
    # ================= TAB 1: MASTER AKUN (COA) =================
    with tab1:
        st.markdown("### Daftar Master Akun & Pengelolaan COA")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_COA}`")
        
        with st.expander("📂 Opsi Lanjutan: Impor / Upload File Excel Master Akun", expanded=False):
            st.write("Unggah file Excel (Kolom A: Nomor Akun, Kolom B: Nama Akun). Kategori terisi otomatis.")
            uploaded_file = st.file_uploader("Pilih file Excel COA (.xlsx)", type=["xlsx"], key="up_coa")
            if uploaded_file is not None:
                try:
                    df_imported = pd.read_excel(uploaded_file, header=0)
                    df_bersih = pd.DataFrame()
                    df_bersih['Kode Akun'] = df_imported.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_bersih['Nama Akun'] = df_imported.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_imported.columns) > 1 else "Tanpa Nama"
                    df_bersih = df_bersih[df_bersih['Kode Akun'] != ""]
                    
                    sub_accs, sub_kats, kats = [], [], []
                    for kode in df_bersih['Kode Akun']:
                        s_acc, s_kat, kat = generate_kategori_otomatis(kode)
                        sub_accs.append(s_acc)
                        sub_kats.append(s_kat)
                        kats.append(kat)
                    df_bersih['Sub Account'] = sub_accs
                    df_bersih['Sub Kategori'] = sub_kats
                    df_bersih['Kategori'] = kats
                    
                    if not df_bersih.empty:
                        st.session_state.master_coa = df_bersih.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                        st.success(f"Berhasil! {len(st.session_state.master_coa)} akun diimpor.")
                        st.rerun()
                except Exception as e:
                    st.error(f"Gagal memproses file: {e}")
            
            if st.button("🗑️ Kosongkan / Reset Tabel COA", key="reset_coa"):
                st.session_state.master_coa = pd.DataFrame(columns=["Kode Akun", "Nama Akun", "Sub Account", "Sub Kategori", "Kategori"])
                simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                st.success("Tabel COA dikosongkan.")
                st.rerun()

        st.markdown("---")
        if not st.session_state.master_coa.empty:
            st.markdown(f"#### 📋 Tabel Master Akun Aktif (Total: {len(st.session_state.master_coa)} Baris Akun)")
            st.dataframe(st.session_state.master_coa, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen ke Excel", use_container_width=True, key="save_coa_perm"):
                simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                st.success("Data berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit Akun Secara Spesifik (Per Nomor Rekening)")
            
            list_kode_akun = st.session_state.master_coa['Kode Akun'].tolist()
            pilih_akun_edit = st.selectbox("Pilih Nomor Akun / Rekening yang Ingin Diedit / Dicek", list_kode_akun, key="select_edit_akun")
            
            if pilih_akun_edit:
                row_akun = st.session_state.master_coa[st.session_state.master_coa['Kode Akun'] == pilih_akun_edit].iloc[0]
                edit_nama_akun = st.text_input("Nama Account (Variabel)", value=row_akun['Nama Akun'], key=f"input_nama_{pilih_akun_edit}")
                
                s_acc_prev, s_kat_prev, kat_prev = generate_kategori_otomatis(pilih_akun_edit)
                st.caption(f"📌 **Auto-Generate Kategori:** Sub Account: `{s_acc_prev}` | Sub Kategori: `{s_kat_prev}` | Kategori: `{kat_prev}`")
                
                col_k1, col_k2 = st.columns(2)
                with col_k1:
                    if st.button("💾 Simpan Perubahan Akun Ini", use_container_width=True, key=f"btn_save_{pilih_akun_edit}"):
                        s_acc, s_kat, kat = generate_kategori_otomatis(pilih_akun_edit)
                        st.session_state.master_coa.loc[st.session_state.master_coa['Kode Akun'] == pilih_akun_edit, ['Nama Akun', 'Sub Account', 'Sub Kategori', 'Kategori']] = [edit_nama_akun, s_acc, s_kat, kat]
                        st.session_state.master_coa = st.session_state.master_coa.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                        st.success(f"Akun **{pilih_akun_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_k2:
                    if st.button("🗑️ Hapus Akun Ini", use_container_width=True, type="secondary", key=f"btn_del_{pilih_akun_edit}"):
                        st.session_state.master_coa = st.session_state.master_coa[st.session_state.master_coa['Kode Akun'] != pilih_akun_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                        st.success(f"Akun **{pilih_akun_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_coa, "Master Kode Rekening (COA)", FILE_MASTER_COA, "coa")

        st.divider()
        with st.form("form_tambah_akun", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Akun Baru")
            c1, c2 = st.columns(2)
            with c1: input_no_akun = st.text_input("Nomor Akun (Cth: 5133.001)")
            with c2: input_nama_akun = st.text_input("Nama Account")
            if st.form_submit_button("➕ Tambahkan Akun"):
                if input_no_akun and input_nama_akun:
                    kode_bersih = input_no_akun.strip()
                    s_acc, s_kat, kat = generate_kategori_otomatis(kode_bersih)
                    data_baru = {"Kode Akun": kode_bersih, "Nama Akun": input_nama_akun, "Sub Account": s_acc, "Sub Kategori": s_kat, "Kategori": kat}
                    st.session_state.master_coa = pd.concat([st.session_state.master_coa, pd.DataFrame([data_baru])], ignore_index=True).drop_duplicates(subset=["Kode Akun"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_coa, FILE_MASTER_COA, 'Master_COA')
                    st.success("Akun berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Nomor dan Nama Akun.")

    # ================= TAB 2: BUSINESS UNIT (BU) =================
    with tab2:
        st.markdown("### Daftar Business Unit (BU) / Proyek")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_BU}`")
        with st.expander("📂 Opsi Lanjutan: Impor File Excel BU", expanded=False):
            up_bu = st.file_uploader("Pilih file Excel BU (.xlsx)", type=["xlsx"], key="up_bu")
            if up_bu:
                try:
                    df_bu = pd.read_excel(up_bu, header=0)
                    df_bu_clean = pd.DataFrame()
                    df_bu_clean['Kode BU'] = df_bu.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_bu_clean['Nama Business Unit'] = df_bu.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_bu.columns) > 1 else ""
                    df_bu_clean['Keterangan'] = df_bu.iloc[:, 2].fillna("").astype(str) if len(df_bu.columns) > 2 else ""
                    df_bu_clean = df_bu_clean[df_bu_clean['Kode BU'] != ""]
                    st.session_state.master_bu = df_bu_clean.reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_bu, FILE_MASTER_BU, 'Master_BU')
                    st.success("Master BU berhasil diimpor!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal: {e}")
        
        if not st.session_state.master_bu.empty:
            st.markdown(f"#### 📋 Tabel Business Unit Aktif (Total: {len(st.session_state.master_bu)} Baris)")
            st.dataframe(st.session_state.master_bu, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen BU ke Excel", use_container_width=True, key="save_bu_perm"):
                simpan_excel_cantik(st.session_state.master_bu, FILE_MASTER_BU, 'Master_BU')
                st.success("Data BU berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit BU Secara Spesifik (Per Kode BU)")
            list_kode_bu = st.session_state.master_bu['Kode BU'].tolist()
            pilih_bu_edit = st.selectbox("Pilih Kode BU yang Ingin Diedit / Dicek", list_kode_bu, key="select_edit_bu")
            
            if pilih_bu_edit:
                row_bu = st.session_state.master_bu[st.session_state.master_bu['Kode BU'] == pilih_bu_edit].iloc[0]
                edit_nama_bu = st.text_input("Nama Business Unit", value=row_bu['Nama Business Unit'], key=f"input_bu_nama_{pilih_bu_edit}")
                edit_ket_bu = st.text_input("Keterangan", value=row_bu.get('Keterangan', ''), key=f"input_bu_ket_{pilih_bu_edit}")
                
                col_b1, col_b2 = st.columns(2)
                with col_b1:
                    if st.button("💾 Simpan Perubahan BU Ini", use_container_width=True, key=f"btn_save_bu_{pilih_bu_edit}"):
                        st.session_state.master_bu.loc[st.session_state.master_bu['Kode BU'] == pilih_bu_edit, ['Nama Business Unit', 'Keterangan']] = [edit_nama_bu, edit_ket_bu]
                        st.session_state.master_bu = st.session_state.master_bu.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_bu, FILE_MASTER_BU, 'Master_BU')
                        st.success(f"BU **{pilih_bu_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_b2:
                    if st.button("🗑️ Hapus BU Ini", use_container_width=True, type="secondary", key=f"btn_del_bu_{pilih_bu_edit}"):
                        st.session_state.master_bu = st.session_state.master_bu[st.session_state.master_bu['Kode BU'] != pilih_bu_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_bu, FILE_MASTER_BU, 'Master_BU')
                        st.success(f"BU **{pilih_bu_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_bu, "Business Unit (BU)", FILE_MASTER_BU, "bu")

        st.divider()
        with st.form("form_tambah_bu", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Business Unit Baru")
            b1, b2, b3 = st.columns(3)
            with b1: k_bu = st.text_input("Kode BU (Cth: BU-01)")
            with b2: n_bu = st.text_input("Nama Business Unit")
            with b3: ket_bu = st.text_input("Keterangan")
            if st.form_submit_button("➕ Tambah BU"):
                if k_bu and n_bu:
                    new_row = {"Kode BU": k_bu.strip(), "Nama Business Unit": n_bu, "Keterangan": ket_bu}
                    st.session_state.master_bu = pd.concat([st.session_state.master_bu, pd.DataFrame([new_row])], ignore_index=True).drop_duplicates(subset=["Kode BU"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_bu, FILE_MASTER_BU, 'Master_BU')
                    st.success("BU berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Kode dan Nama BU.")

    # ================= TAB 3: NAMA PELANGGAN =================
    with tab3:
        st.markdown("### Daftar Nama Pelanggan (Customer)")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_PELANGGAN}`")
        with st.expander("📂 Opsi Lanjutan: Impor File Excel Pelanggan", expanded=False):
            up_pel = st.file_uploader("Pilih file Excel Pelanggan (.xlsx)", type=["xlsx"], key="up_pel")
            if up_pel:
                try:
                    df_pel = pd.read_excel(up_pel, header=0)
                    df_pel_clean = pd.DataFrame()
                    df_pel_clean['Kode Pelanggan'] = df_pel.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_pel_clean['Nama Pelanggan'] = df_pel.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_pel.columns) > 1 else ""
                    df_pel_clean['Alamat'] = df_pel.iloc[:, 2].fillna("").astype(str) if len(df_pel.columns) > 2 else ""
                    df_pel_clean['Kontak'] = df_pel.iloc[:, 3].fillna("").astype(str) if len(df_pel.columns) > 3 else ""
                    df_pel_clean = df_pel_clean[df_pel_clean['Kode Pelanggan'] != ""]
                    st.session_state.master_pelanggan = df_pel_clean.reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_pelanggan, FILE_MASTER_PELANGGAN, 'Master_Pelanggan')
                    st.success("Master Pelanggan berhasil diimpor!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal: {e}")

        if not st.session_state.master_pelanggan.empty:
            st.markdown(f"#### 📋 Tabel Pelanggan Aktif (Total: {len(st.session_state.master_pelanggan)} Baris)")
            st.dataframe(st.session_state.master_pelanggan, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen Pelanggan ke Excel", use_container_width=True, key="save_pel_perm"):
                simpan_excel_cantik(st.session_state.master_pelanggan, FILE_MASTER_PELANGGAN, 'Master_Pelanggan')
                st.success("Data Pelanggan berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit Pelanggan Secara Spesifik (Per Kode Pelanggan)")
            list_kode_pel = st.session_state.master_pelanggan['Kode Pelanggan'].tolist()
            pilih_pel_edit = st.selectbox("Pilih Kode Pelanggan yang Ingin Diedit / Dicek", list_kode_pel, key="select_edit_pel")
            
            if pilih_pel_edit:
                row_pel = st.session_state.master_pelanggan[st.session_state.master_pelanggan['Kode Pelanggan'] == pilih_pel_edit].iloc[0]
                edit_nama_pel = st.text_input("Nama Pelanggan", value=row_pel['Nama Pelanggan'], key=f"input_pel_nama_{pilih_pel_edit}")
                edit_alamat_pel = st.text_input("Alamat", value=row_pel.get('Alamat', ''), key=f"input_pel_alamat_{pilih_pel_edit}")
                edit_kontak_pel = st.text_input("Kontak / No Telp", value=row_pel.get('Kontak', ''), key=f"input_pel_kontak_{pilih_pel_edit}")
                
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    if st.button("💾 Simpan Perubahan Pelanggan Ini", use_container_width=True, key=f"btn_save_pel_{pilih_pel_edit}"):
                        st.session_state.master_pelanggan.loc[st.session_state.master_pelanggan['Kode Pelanggan'] == pilih_pel_edit, ['Nama Pelanggan', 'Alamat', 'Kontak']] = [edit_nama_pel, edit_alamat_pel, edit_kontak_pel]
                        st.session_state.master_pelanggan = st.session_state.master_pelanggan.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_pelanggan, FILE_MASTER_PELANGGAN, 'Master_Pelanggan')
                        st.success(f"Pelanggan **{pilih_pel_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_p2:
                    if st.button("🗑️ Hapus Pelanggan Ini", use_container_width=True, type="secondary", key=f"btn_del_pel_{pilih_pel_edit}"):
                        st.session_state.master_pelanggan = st.session_state.master_pelanggan[st.session_state.master_pelanggan['Kode Pelanggan'] != pilih_pel_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_pelanggan, FILE_MASTER_PELANGGAN, 'Master_Pelanggan')
                        st.success(f"Pelanggan **{pilih_pel_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_pelanggan, "Daftar Pelanggan", FILE_MASTER_PELANGGAN, "pelanggan")

        st.divider()
        with st.form("form_tambah_pelanggan", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Pelanggan Baru")
            p1, p2 = st.columns(2)
            with p1: 
                kp = st.text_input("Kode Pelanggan (Cth: CUS-01)")
                alamat_p = st.text_input("Alamat")
            with p2: 
                np = st.text_input("Nama Pelanggan")
                kontak_p = st.text_input("Kontak / No Telp")
            if st.form_submit_button("➕ Tambah Pelanggan"):
                if kp and np:
                    new_row = {"Kode Pelanggan": kp.strip(), "Nama Pelanggan": np, "Alamat": alamat_p, "Kontak": kontak_p}
                    st.session_state.master_pelanggan = pd.concat([st.session_state.master_pelanggan, pd.DataFrame([new_row])], ignore_index=True).drop_duplicates(subset=["Kode Pelanggan"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_pelanggan, FILE_MASTER_PELANGGAN, 'Master_Pelanggan')
                    st.success("Pelanggan berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Kode dan Nama Pelanggan.")

    # ================= TAB 4: NAMA PEMASOK =================
    with tab4:
        st.markdown("### Daftar Nama Pemasok (Supplier)")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_PEMASOK}`")
        with st.expander("📂 Opsi Lanjutan: Impor File Excel Pemasok", expanded=False):
            up_pem = st.file_uploader("Pilih file Excel Pemasok (.xlsx)", type=["xlsx"], key="up_pem")
            if up_pem:
                try:
                    df_pem = pd.read_excel(up_pem, header=0)
                    df_pem_clean = pd.DataFrame()
                    df_pem_clean['Kode Pemasok'] = df_pem.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_pem_clean['Nama Pemasok'] = df_pem.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_pem.columns) > 1 else ""
                    df_pem_clean['Alamat'] = df_pem.iloc[:, 2].fillna("").astype(str) if len(df_pem.columns) > 2 else ""
                    df_pem_clean['Kontak'] = df_pem.iloc[:, 3].fillna("").astype(str) if len(df_pem.columns) > 3 else ""
                    df_pem_clean = df_pem_clean[df_pem_clean['Kode Pemasok'] != ""]
                    st.session_state.master_pemasok = df_pem_clean.reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_pemasok, FILE_MASTER_PEMASOK, 'Master_Pemasok')
                    st.success("Master Pemasok berhasil diimpor!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal: {e}")

        if not st.session_state.master_pemasok.empty:
            st.markdown(f"#### 📋 Tabel Pemasok Aktif (Total: {len(st.session_state.master_pemasok)} Baris)")
            st.dataframe(st.session_state.master_pemasok, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen Pemasok ke Excel", use_container_width=True, key="save_pem_perm"):
                simpan_excel_cantik(st.session_state.master_pemasok, FILE_MASTER_PEMASOK, 'Master_Pemasok')
                st.success("Data Pemasok berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit Pemasok Secara Spesifik (Per Kode Pemasok)")
            list_kode_pem = st.session_state.master_pemasok['Kode Pemasok'].tolist()
            pilih_pem_edit = st.selectbox("Pilih Kode Pemasok yang Ingin Diedit / Dicek", list_kode_pem, key="select_edit_pem")
            
            if pilih_pem_edit:
                row_pem = st.session_state.master_pemasok[st.session_state.master_pemasok['Kode Pemasok'] == pilih_pem_edit].iloc[0]
                edit_nama_pem = st.text_input("Nama Pemasok", value=row_pem['Nama Pemasok'], key=f"input_pem_nama_{pilih_pem_edit}")
                edit_alamat_pem = st.text_input("Alamat", value=row_pem.get('Alamat', ''), key=f"input_pem_alamat_{pilih_pem_edit}")
                edit_kontak_pem = st.text_input("Kontak", value=row_pem.get('Kontak', ''), key=f"input_pem_kontak_{pilih_pem_edit}")
                
                col_pm1, col_pm2 = st.columns(2)
                with col_pm1:
                    if st.button("💾 Simpan Perubahan Pemasok Ini", use_container_width=True, key=f"btn_save_pem_{pilih_pem_edit}"):
                        st.session_state.master_pemasok.loc[st.session_state.master_pemasok['Kode Pemasok'] == pilih_pem_edit, ['Nama Pemasok', 'Alamat', 'Kontak']] = [edit_nama_pem, edit_alamat_pem, edit_kontak_pem]
                        st.session_state.master_pemasok = st.session_state.master_pemasok.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_pemasok, FILE_MASTER_PEMASOK, 'Master_Pemasok')
                        st.success(f"Pemasok **{pilih_pem_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_pm2:
                    if st.button("🗑️ Hapus Pemasok Ini", use_container_width=True, type="secondary", key=f"btn_del_pem_{pilih_pem_edit}"):
                        st.session_state.master_pemasok = st.session_state.master_pemasok[st.session_state.master_pemasok['Kode Pemasok'] != pilih_pem_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_pemasok, FILE_MASTER_PEMASOK, 'Master_Pemasok')
                        st.success(f"Pemasok **{pilih_pem_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_pemasok, "Daftar Pemasok", FILE_MASTER_PEMASOK, "pemasok")

        st.divider()
        with st.form("form_tambah_pemasok", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Pemasok Baru")
            pm1, pm2 = st.columns(2)
            with pm1: 
                kpm = st.text_input("Kode Pemasok (Cth: SUP-01)")
                alamat_pm = st.text_input("Alamat Pemasok")
            with pm2: 
                npm = st.text_input("Nama Pemasok")
                kontak_pm = st.text_input("Kontak / No Telp")
            if st.form_submit_button("➕ Tambah Pemasok"):
                if kpm and npm:
                    new_row = {"Kode Pemasok": kpm.strip(), "Nama Pemasok": npm, "Alamat": alamat_pm, "Kontak": kontak_pm}
                    st.session_state.master_pemasok = pd.concat([st.session_state.master_pemasok, pd.DataFrame([new_row])], ignore_index=True).drop_duplicates(subset=["Kode Pemasok"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_pemasok, FILE_MASTER_PEMASOK, 'Master_Pemasok')
                    st.success("Pemasok berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Kode dan Nama Pemasok.")

    # ================= TAB 5: GUDANG =================
    with tab5:
        st.markdown("### Daftar Gudang (Warehouse)")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_GUDANG}`")
        with st.expander("📂 Opsi Lanjutan: Impor File Excel Gudang", expanded=False):
            up_gud = st.file_uploader("Pilih file Excel Gudang (.xlsx)", type=["xlsx"], key="up_gud")
            if up_gud:
                try:
                    df_gud = pd.read_excel(up_gud, header=0)
                    df_gud_clean = pd.DataFrame()
                    df_gud_clean['Kode Gudang'] = df_gud.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_gud_clean['Nama Gudang'] = df_gud.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_gud.columns) > 1 else ""
                    df_gud_clean['Lokasi'] = df_gud.iloc[:, 2].fillna("").astype(str) if len(df_gud.columns) > 2 else ""
                    df_gud_clean['Keterangan'] = df_gud.iloc[:, 3].fillna("").astype(str) if len(df_gud.columns) > 3 else ""
                    df_gud_clean = df_gud_clean[df_gud_clean['Kode Gudang'] != ""]
                    st.session_state.master_gudang = df_gud_clean.reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_gudang, FILE_MASTER_GUDANG, 'Master_Gudang')
                    st.success("Master Gudang berhasil diimpor!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal: {e}")

        if not st.session_state.master_gudang.empty:
            st.markdown(f"#### 📋 Tabel Gudang Aktif (Total: {len(st.session_state.master_gudang)} Baris)")
            st.dataframe(st.session_state.master_gudang, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen Gudang ke Excel", use_container_width=True, key="save_gud_perm"):
                simpan_excel_cantik(st.session_state.master_gudang, FILE_MASTER_GUDANG, 'Master_Gudang')
                st.success("Data Gudang berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit Gudang Secara Spesifik (Per Kode Gudang)")
            list_kode_gud = st.session_state.master_gudang['Kode Gudang'].tolist()
            pilih_gud_edit = st.selectbox("Pilih Kode Gudang yang Ingin Diedit / Dicek", list_kode_gud, key="select_edit_gud")
            
            if pilih_gud_edit:
                row_gud = st.session_state.master_gudang[st.session_state.master_gudang['Kode Gudang'] == pilih_gud_edit].iloc[0]
                edit_nama_gud = st.text_input("Nama Gudang", value=row_gud['Nama Gudang'], key=f"input_gud_nama_{pilih_gud_edit}")
                edit_lokasi_gud = st.text_input("Lokasi", value=row_gud.get('Lokasi', ''), key=f"input_gud_lokasi_{pilih_gud_edit}")
                edit_ket_gud = st.text_input("Keterangan", value=row_gud.get('Keterangan', ''), key=f"input_gud_ket_{pilih_gud_edit}")
                
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    if st.button("💾 Simpan Perubahan Gudang Ini", use_container_width=True, key=f"btn_save_gud_{pilih_gud_edit}"):
                        st.session_state.master_gudang.loc[st.session_state.master_gudang['Kode Gudang'] == pilih_gud_edit, ['Nama Gudang', 'Lokasi', 'Keterangan']] = [edit_nama_gud, edit_lokasi_gud, edit_ket_gud]
                        st.session_state.master_gudang = st.session_state.master_gudang.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_gudang, FILE_MASTER_GUDANG, 'Master_Gudang')
                        st.success(f"Gudang **{pilih_gud_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_g2:
                    if st.button("🗑️ Hapus Gudang Ini", use_container_width=True, type="secondary", key=f"btn_del_gud_{pilih_gud_edit}"):
                        st.session_state.master_gudang = st.session_state.master_gudang[st.session_state.master_gudang['Kode Gudang'] != pilih_gud_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_gudang, FILE_MASTER_GUDANG, 'Master_Gudang')
                        st.success(f"Gudang **{pilih_gud_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_gudang, "Daftar Gudang", FILE_MASTER_GUDANG, "gudang")

        st.divider()
        with st.form("form_tambah_gudang", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Gudang Baru")
            g1, g2 = st.columns(2)
            with g1: 
                kg = st.text_input("Kode Gudang (Cth: WH-01)")
                lokasi_g = st.text_input("Lokasi Gudang")
            with g2: 
                ng = st.text_input("Nama Gudang")
                ket_g = st.text_input("Keterangan Gudang")
            if st.form_submit_button("➕ Tambah Gudang"):
                if kg and ng:
                    new_row = {"Kode Gudang": kg.strip(), "Nama Gudang": ng, "Lokasi": lokasi_g, "Keterangan": ket_g}
                    st.session_state.master_gudang = pd.concat([st.session_state.master_gudang, pd.DataFrame([new_row])], ignore_index=True).drop_duplicates(subset=["Kode Gudang"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_gudang, FILE_MASTER_GUDANG, 'Master_Gudang')
                    st.success("Gudang berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Kode dan Nama Gudang.")

    # ================= TAB 6: ALOKASI ALAT / UNIT =================
    with tab6:
        st.markdown("### Daftar Alokasi Alat / Unit")
        st.info(f"📂 Direktori File Excel: folder utama aplikasi -> `{FILE_MASTER_ALAT}`")
        with st.expander("📂 Opsi Lanjutan: Impor File Excel Alat / Unit", expanded=False):
            up_alt = st.file_uploader("Pilih file Excel Alat (.xlsx)", type=["xlsx"], key="up_alt")
            if up_alt:
                try:
                    df_alt = pd.read_excel(up_alt, header=0)
                    df_alt_clean = pd.DataFrame()
                    df_alt_clean['Kode Alat'] = df_alt.iloc[:, 0].fillna("").astype(str).str.strip()
                    df_alt_clean['Nama Alat / Unit'] = df_alt.iloc[:, 1].fillna("Tanpa Nama").astype(str) if len(df_alt.columns) > 1 else ""
                    df_alt_clean['Jenis / Kategori'] = df_alt.iloc[:, 2].fillna("").astype(str) if len(df_alt.columns) > 2 else ""
                    df_alt_clean['No. Polisi / Serial'] = df_alt.iloc[:, 3].fillna("").astype(str) if len(df_alt.columns) > 3 else ""
                    df_alt_clean['Keterangan'] = df_alt.iloc[:, 4].fillna("").astype(str) if len(df_alt.columns) > 4 else ""
                    df_alt_clean = df_alt_clean[df_alt_clean['Kode Alat'] != ""]
                    st.session_state.master_alat = df_alt_clean.reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_alat, FILE_MASTER_ALAT, 'Master_Alat')
                    st.success("Master Alat / Unit berhasil diimpor!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal: {e}")

        if not st.session_state.master_alat.empty:
            st.markdown(f"#### 📋 Tabel Alokasi Alat Aktif (Total: {len(st.session_state.master_alat)} Baris)")
            st.dataframe(st.session_state.master_alat, use_container_width=True, height=450)
            
            if st.button("💾 Simpan Permanen Alat ke Excel", use_container_width=True, key="save_alt_perm"):
                simpan_excel_cantik(st.session_state.master_alat, FILE_MASTER_ALAT, 'Master_Alat')
                st.success("Data Alat berhasil disimpan permanen ke Excel!")

            st.markdown("---")
            st.markdown("#### 🔄 Panggil Ulang & Edit Alat Secara Spesifik (Per Kode Alat)")
            list_kode_alt = st.session_state.master_alat['Kode Alat'].tolist()
            pilih_alt_edit = st.selectbox("Pilih Kode Alat yang Ingin Diedit / Dicek", list_kode_alt, key="select_edit_alt")
            
            if pilih_alt_edit:
                row_alt = st.session_state.master_alat[st.session_state.master_alat['Kode Alat'] == pilih_alt_edit].iloc[0]
                edit_nama_alt = st.text_input("Nama Alat / Unit", value=row_alt['Nama Alat / Unit'], key=f"input_alt_nama_{pilih_alt_edit}")
                edit_jenis_alt = st.text_input("Jenis / Kategori", value=row_alt.get('Jenis / Kategori', ''), key=f"input_alt_jenis_{pilih_alt_edit}")
                edit_nopol_alt = st.text_input("No. Polisi / Serial", value=row_alt.get('No. Polisi / Serial', ''), key=f"input_alt_nopol_{pilih_alt_edit}")
                edit_ket_alt = st.text_input("Keterangan", value=row_alt.get('Keterangan', ''), key=f"input_alt_ket_{pilih_alt_edit}")
                
                col_a1, col_a2 = st.columns(2)
                with col_a1:
                    if st.button("💾 Simpan Perubahan Alat Ini", use_container_width=True, key=f"btn_save_alt_{pilih_alt_edit}"):
                        st.session_state.master_alat.loc[st.session_state.master_alat['Kode Alat'] == pilih_alt_edit, ['Nama Alat / Unit', 'Jenis / Kategori', 'No. Polisi / Serial', 'Keterangan']] = [edit_nama_alt, edit_jenis_alt, edit_nopol_alt, edit_ket_alt]
                        st.session_state.master_alat = st.session_state.master_alat.reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_alat, FILE_MASTER_ALAT, 'Master_Alat')
                        st.success(f"Alat **{pilih_alt_edit}** berhasil diperbarui!")
                        st.rerun()
                with col_a2:
                    if st.button("🗑️ Hapus Alat Ini", use_container_width=True, type="secondary", key=f"btn_del_alt_{pilih_alt_edit}"):
                        st.session_state.master_alat = st.session_state.master_alat[st.session_state.master_alat['Kode Alat'] != pilih_alt_edit].reset_index(drop=True)
                        simpan_excel_cantik(st.session_state.master_alat, FILE_MASTER_ALAT, 'Master_Alat')
                        st.success(f"Alat **{pilih_alt_edit}** berhasil dihapus!")
                        st.rerun()

            render_preview_dan_cetak(st.session_state.master_alat, "Alokasi Alat / Unit", FILE_MASTER_ALAT, "alat")

        st.divider()
        with st.form("form_tambah_alat", clear_on_submit=True):
            st.markdown("#### ➕ Tambah Alat / Unit Baru")
            a1, a2 = st.columns(2)
            with a1: 
                ka = st.text_input("Kode Alat (Cth: TR-01)")
                jenis_a = st.text_input("Jenis / Kategori Alat")
                ket_a = st.text_input("Keterangan Tambahan")
            with a2: 
                na = st.text_input("Nama Alat / Unit")
                nopol_a = st.text_input("No. Polisi / Serial")
            if st.form_submit_button("➕ Tambah Alat / Unit"):
                if ka and na:
                    new_row = {
                        "Kode Alat": ka.strip(), 
                        "Nama Alat / Unit": na, 
                        "Jenis / Kategori": jenis_a, 
                        "No. Polisi / Serial": nopol_a,
                        "Keterangan": ket_a
                    }
                    st.session_state.master_alat = pd.concat([st.session_state.master_alat, pd.DataFrame([new_row])], ignore_index=True).drop_duplicates(subset=["Kode Alat"], keep="last").reset_index(drop=True)
                    simpan_excel_cantik(st.session_state.master_alat, FILE_MASTER_ALAT, 'Master_Alat')
                    st.success("Alat / Unit berhasil ditambahkan!")
                    st.rerun()
                else:
                    st.error("Lengkapi Kode Alat dan Nama Alat / Unit.")